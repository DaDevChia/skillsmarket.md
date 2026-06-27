"""SkillsFuture / MySkillsFuture course directory ingestion.

Downloads the official MySkillsFuture Course Directory (data.gov.sg dataset
d_b5802b76f409764c16dde4bf2feb19cd), then builds a compact per-skill index of
real courses (title, provider, fee, hours, and a real course-detail link). This
is real SkillsFuture data, cached to disk; matches are by title/keyword.
"""

from __future__ import annotations

import io
import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import httpx

from skillsmarket.courses import COURSE_DIRECTORY_DATASET_ID

POLL_URL = f"https://api-open.data.gov.sg/v1/public/api/datasets/{COURSE_DIRECTORY_DATASET_ID}/poll-download"
COURSE_DETAIL_URL = (
    "https://www.myskillsfuture.gov.sg/content/portal/en/training-exchange/"
    "course-directory/course-detail.html?courseReferenceNumber="
)
COURSES_PATH = Path("data/skillsfuture_courses.json")


def course_detail_url(reference: str) -> str:
    return f"{COURSE_DETAIL_URL}{reference}"


def download_course_rows() -> list[dict[str, Any]]:
    poll = httpx.get(POLL_URL, timeout=30).json()
    url = (poll.get("data") or {}).get("url")
    if not url:
        raise RuntimeError("data.gov.sg poll-download returned no URL")
    raw = httpx.get(url, timeout=120).content

    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(raw), read_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    header = [str(cell) for cell in next(rows)]
    courses: list[dict[str, Any]] = []
    for values in rows:
        record = dict(zip(header, values))
        title = record.get("coursetitle")
        if not title:
            continue
        courses.append(
            {
                "ref": record.get("coursereferencenumber"),
                "title": str(title),
                "provider": record.get("trainingprovideralias"),
                "fee": record.get("course_fee_after_subsidies") or record.get("full_course_fee"),
                "hours": record.get("number_of_hours"),
                "learn": str(record.get("what_you_learn") or "")[:200],
            }
        )
    workbook.close()
    return courses


def build_course_index(skill_names: list[str], courses: list[dict[str, Any]], per_skill: int = 4) -> dict[str, Any]:
    lowered = [(course, course["title"].lower(), course["learn"].lower()) for course in courses]
    index: dict[str, list[dict[str, Any]]] = {}
    for name in skill_names:
        needle = name.lower()
        matches: list[dict[str, Any]] = []
        for course, title_l, learn_l in lowered:
            if needle in title_l or needle in learn_l:
                matches.append(
                    {
                        "ref": course["ref"],
                        "title": course["title"],
                        "provider": course["provider"],
                        "fee": course["fee"],
                        "hours": course["hours"],
                        "url": course_detail_url(course["ref"]) if course["ref"] else None,
                    }
                )
                if len(matches) >= per_skill:
                    break
        if matches:
            index[needle] = matches
    return {
        "fetched_at": datetime.now(timezone.utc).isoformat(),
        "source": "MySkillsFuture Course Directory (data.gov.sg)",
        "dataset_id": COURSE_DIRECTORY_DATASET_ID,
        "course_count": len(courses),
        "skills": index,
    }


def write_courses(index: dict[str, Any]) -> Path:
    COURSES_PATH.parent.mkdir(parents=True, exist_ok=True)
    COURSES_PATH.write_text(json.dumps(index))
    return COURSES_PATH


def load_courses() -> dict[str, Any] | None:
    if not COURSES_PATH.exists():
        return None
    try:
        return json.loads(COURSES_PATH.read_text())
    except (json.JSONDecodeError, OSError):
        return None


def courses_for_skill(name: str, courses: dict[str, Any] | None = None) -> dict[str, Any]:
    courses = courses or load_courses()
    if not courses:
        return {"found": False, "matches": []}
    matches = courses.get("skills", {}).get(name.lower(), [])
    return {
        "found": bool(matches),
        "source": courses.get("source"),
        "fetched_at": courses.get("fetched_at"),
        "matches": matches,
    }
